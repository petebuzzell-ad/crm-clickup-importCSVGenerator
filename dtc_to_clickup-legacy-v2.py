#!/usr/bin/env python3
"""
DTC Calendar to ClickUp CSV Converter

Converts DTC Calendar Excel files for Arcadia Digital brands (PB, TGW) into ClickUp-importable CSV format.

================================================================================
INTELLECTUAL PROPERTY NOTICE
================================================================================

This tool is part of Arcadia Digital's proprietary operational infrastructure.

This material is confidential and may not be shared, reproduced, or used 
outside of Arcadia Digital without explicit authorization.

Unauthorized distribution, modification, or use of this software is prohibited.

© 2026 Arcadia Digital. All rights reserved.

================================================================================
"""

import sys
import csv
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class DTCtoClickUpConverter:
    """Converts DTC Calendar Excel files to ClickUp CSV format."""

    # Priority mapping for campaign types
    CAMPAIGN_TYPE_PRIORITY = {
        'Product Launches': 'High',
        'Promotions': 'High',
        'Story Telling': 'Normal',
        'Brand Moments': 'Normal',
        'Problem Solving': 'Normal',
        'Promo': 'High',
    }

    # Priority mapping for product launch priorities
    PRODUCT_PRIORITY_MAP = {
        'A': 'Urgent',
        'B': 'High',
        'C': 'Normal',
        'D': 'Low',
    }

    def __init__(self, excel_file: str, brand: str, output_file: str):
        """
        Initialize converter.

        Args:
            excel_file: Path to input Excel file
            brand: Brand name (PB or TGW)
            output_file: Path to output CSV file
        """
        self.excel_file = excel_file
        self.brand = brand
        self.output_file = output_file
        self.workbook = None
        self.tasks: List[Dict] = []
        self.stats = {
            'product_launches': 0,
            'campaign_tasks': 0,
            'sheets_processed': 0,
        }

    def load_workbook_safe(self) -> bool:
        """
        Load Excel workbook with error handling.

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            self.workbook = load_workbook(self.excel_file, data_only=True)
            print(f"Loaded workbook: {self.excel_file}")
            return True
        except FileNotFoundError:
            print(f"Error: File not found - {self.excel_file}", file=sys.stderr)
            return False
        except Exception as e:
            print(f"Error loading workbook: {e}", file=sys.stderr)
            return False

    def get_cell_value(self, sheet, row: int, col: int) -> Optional[str]:
        """
        Get cell value safely, handling None and converting to string.

        Args:
            sheet: openpyxl worksheet
            row: Row number (1-indexed)
            col: Column number (1-indexed)

        Returns:
            Cleaned string value or None
        """
        cell = sheet.cell(row, col)
        value = cell.value
        if value is None:
            return None
        # Convert to string and strip whitespace
        value_str = str(value).strip()
        return value_str if value_str else None

    def parse_date(self, date_str: str) -> Optional[str]:
        """
        Parse various date formats and return MM/DD/YYYY format.

        Args:
            date_str: Date string in various formats

        Returns:
            Formatted date string MM/DD/YYYY or None
        """
        if not date_str:
            return None

        date_str = str(date_str).strip()

        # Try common formats
        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%m/%d/%Y',
            '%m/%d/%Y %H:%M',
            '%m/%d/%Y %I:%M %p',
            '%Y-%m-%d',
            '%m-%d-%Y',
        ]

        for fmt in formats:
            try:
                dt = datetime.strptime(date_str, fmt)
                return dt.strftime('%m/%d/%Y')
            except ValueError:
                continue

        # If datetime object, format it
        if isinstance(date_str, datetime):
            return date_str.strftime('%m/%d/%Y')

        return None

    def due_date_two_weeks_before(self, send_date_str: str) -> Optional[str]:
        """
        Given a date string (MM/DD/YYYY), return the date 14 days earlier in the same format.

        Args:
            send_date_str: Date string in MM/DD/YYYY format

        Returns:
            Formatted date string MM/DD/YYYY or None if parsing fails
        """
        if not send_date_str:
            return None
        parsed = self.parse_date(send_date_str)
        if not parsed:
            return None
        try:
            dt = datetime.strptime(parsed, '%m/%d/%Y')
            earlier = dt - timedelta(days=14)
            return earlier.strftime('%m/%d/%Y')
        except ValueError:
            return None

    def parse_due_date_from_header(self, date_str: str) -> Optional[str]:
        """
        Parse due date from header like 'DUE 2/3 10 AM CT'.

        Args:
            date_str: Header string containing due date

        Returns:
            Formatted date string MM/DD/YYYY or None
        """
        if not date_str:
            return None

        # Extract date pattern like "2/3" or "2/3/24"
        match = re.search(r'(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?', date_str)
        if match:
            month, day = match.group(1), match.group(2)
            year = match.group(3) if match.group(3) else str(datetime.now().year)

            # Handle 2-digit year
            if len(year) == 2:
                year = f"20{year}"

            try:
                dt = datetime(int(year), int(month), int(day))
                return dt.strftime('%m/%d/%Y')
            except ValueError:
                return None

        return None

    def clean_text(self, text: str) -> str:
        """
        Clean text: strip whitespace and normalize line breaks.

        Args:
            text: Text to clean

        Returns:
            Cleaned text
        """
        if not text:
            return ''
        return re.sub(r'\s+', ' ', str(text).strip())

    def normalize_multiline_text(self, text: str) -> str:
        """
        Normalize multiline text for CSV.

        Args:
            text: Text potentially containing line breaks

        Returns:
            Text with normalized line breaks
        """
        if not text:
            return ''
        # Replace multiple spaces with single space, preserve intentional breaks
        return re.sub(r'  +', ' ', str(text).strip())

    def extract_product_launch_tasks(self) -> int:
        """
        Extract product launch tasks from 'Product Launch Calendar' sheet.
        Handles both PB format (6 cols) and TGW format (8 cols with SKET Task, PAR Date, PO#).
        """
        sheet_name = 'Product Launch Calendar'

        if sheet_name not in self.workbook.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found")
            return 0

        sheet = self.workbook[sheet_name]
        count = 0

        # Auto-detect column layout from header row
        headers = {}
        for col in range(1, min(sheet.max_column + 1, 15)):
            header_val = self.get_cell_value(sheet, 1, col)
            if header_val:
                headers[col] = header_val.lower()

        # Find column indices
        desc_col = launch_col = priority_col = None
        for col_idx, header in headers.items():
            if 'description' in header:
                desc_col = col_idx
            elif 'launch date' in header:
                launch_col = col_idx
            elif 'priority' in header:
                priority_col = col_idx

        if not desc_col or not launch_col:
            print(f"Warning: Could not find required columns in '{sheet_name}'")
            return count

        # Extract rows
        for row in range(2, sheet.max_row + 1):
            desc = self.get_cell_value(sheet, row, desc_col)
            if not desc:
                continue

            launch_date_raw = self.get_cell_value(sheet, row, launch_col)
            launch_date = self.parse_date(launch_date_raw) if launch_date_raw else None

            priority_raw = self.get_cell_value(sheet, row, priority_col) if priority_col else None
            priority = self.PRODUCT_PRIORITY_MAP.get(priority_raw, 'Normal') if priority_raw else 'Normal'

            task_name = f"[Product Launch] {desc}"

            task = {
                'Task Name': task_name,
                'Task Description': f"Product: {desc}\nLaunch Date: {launch_date or 'TBD'}",
                'Due Date': launch_date or '',
                'Start Date': '',
                'Priority': priority,
                'Status': 'Open',
                'Tags': f'{self.brand}, Product Launch',
            }

            self.tasks.append(task)
            count += 1

        self.stats['product_launches'] = count
        return count

    def _scan_for_label(self, sheet, label_text: str, start_row: int, end_row: int) -> Optional[int]:
        """
        Scan column B for a label between start_row and end_row.

        Args:
            sheet: openpyxl worksheet
            label_text: Text to search for (case-insensitive)
            start_row: Starting row number
            end_row: Ending row number

        Returns:
            Row number if found, None otherwise
        """
        for row in range(start_row, end_row + 1):
            val = self.get_cell_value(sheet, row, 2)  # Column B
            if val and label_text.lower() in val.lower():
                return row
        return None

    def extract_campaign_tasks_from_sheet(self, sheet_name: str) -> int:
        """
        Extract campaign tasks from a weekly sheet (Wk6, Wk7, etc.).
        Creates email brief tasks with comprehensive marketing/merchandising details.

        Returns:
            Number of tasks extracted
        """
        if sheet_name not in self.workbook.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found")
            return 0

        sheet = self.workbook[sheet_name]
        count = 0

        # Extract week number from sheet name
        week_match = re.search(r'[Ww]k\s*(\d+)', sheet_name)
        week_num = f"Wk{week_match.group(1)}" if week_match else sheet_name

        # Find the due date from header (typically in first few rows)
        due_date = None
        for row in range(1, min(5, sheet.max_row + 1)):
            for col in range(1, min(5, sheet.max_column + 1)):
                cell_val = self.get_cell_value(sheet, row, col)
                if cell_val and 'DUE' in cell_val.upper():
                    due_date = self.parse_due_date_from_header(cell_val)
                    if due_date:
                        break
            if due_date:
                break

        # Find key rows in column B
        def find_row(label: str) -> Optional[int]:
            return self._scan_for_label(sheet, label, 1, min(sheet.max_row, 50))

        row_campaign_type = find_row('Campaign Type')
        row_campaign_name = find_row('Campaign Name')
        row_overview = find_row('Email Overview')
        row_date_of_send = find_row('Date of Send')
        row_time_of_send = find_row('Time of Send')

        # Iterate through columns (C onwards) to find campaigns
        for col in range(3, min(sheet.max_column + 1, 50)):
            # Check if this column has a campaign
            campaign_type = None
            campaign_name = None

            if row_campaign_type:
                campaign_type = self.get_cell_value(sheet, row_campaign_type, col)
            if row_campaign_name:
                campaign_name = self.get_cell_value(sheet, row_campaign_name, col)

            # Skip if no campaign name
            if not campaign_name or campaign_name.upper() in ['TBD', 'N/A', '']:
                continue

            # Get send date info
            start_date = None
            time_of_send = None
            day_of_week = None

            if row_date_of_send:
                date_val = self.get_cell_value(sheet, row_date_of_send, col)
                start_date = self.parse_date(date_val) if date_val else None
                if start_date:
                    try:
                        dt = datetime.strptime(start_date, '%m/%d/%Y')
                        day_of_week = dt.strftime('%A')
                    except:
                        pass

            if row_time_of_send:
                time_of_send = self.get_cell_value(sheet, row_time_of_send, col)

            # Determine priority
            priority = self.CAMPAIGN_TYPE_PRIORITY.get(campaign_type, 'Normal') if campaign_type else 'Normal'

            # Build task name
            task_name = f"[Email] {campaign_name}"

            # Build task description sections
            description_parts = []
            description_parts.append("== EMAIL BRIEF ==")
            description_parts.append(f"Send Date: {start_date or 'TBD'} ({day_of_week or ''}) {time_of_send or ''}")
            description_parts.append(f"Campaign Type: {campaign_type or 'TBD'}")

            # Email content section
            if row_overview:
                overview = self.get_cell_value(sheet, row_overview, col)
                if overview:
                    description_parts.append(f"Overview: {overview}")

            # Copy requirements
            copy_row = find_row('Required Copy')
            if copy_row:
                copy_val = self.get_cell_value(sheet, copy_row, col)
                if copy_val:
                    description_parts.append(f"Required Copy: {copy_val}")

            # Creative assets
            assets_row = find_row('Creative Assets')
            if assets_row:
                assets = self.get_cell_value(sheet, assets_row, col)
                if assets:
                    description_parts.append(f"Creative Assets: {assets}")

            # Promotion details (if applicable)
            promo_parts = []
            discount_row = find_row('Discount')
            code_row = find_row('Promo Code')
            terms_row = find_row('Terms')

            if discount_row:
                dv = self.get_cell_value(sheet, discount_row, col)
                if dv:
                    promo_parts.append(f"Discount: {dv}")
            if code_row:
                cv = self.get_cell_value(sheet, code_row, col)
                if cv:
                    promo_parts.append(f"Promo Code: {cv}")
            if terms_row:
                tv = self.get_cell_value(sheet, terms_row, col)
                if tv:
                    promo_parts.append(f"Terms: {tv}")

            if promo_parts:
                description_parts.append(f"--- Promotion ---")
                description_parts.extend(promo_parts)

            # Merchandising section - DAM, Hero, Landing Page, Featured Products
            description_parts.append(f"--- Merchandising ---")

            dam_row = find_row('DAM Assets')
            hero_row = find_row('Hero Product URL')
            special_row = find_row('makes this product special')
            inventory_row = find_row('Inventory In House')
            landing_row = find_row('Landing Page')

            if dam_row:
                dv = self.get_cell_value(sheet, dam_row, col)
                if dv:
                    description_parts.append(f"DAM Assets: {dv}")
            if hero_row:
                hv = self.get_cell_value(sheet, hero_row, col)
                if hv:
                    description_parts.append(f"Hero Product URL: {hv}")
            if special_row:
                sv = self.get_cell_value(sheet, special_row, col)
                if sv:
                    description_parts.append(f"What Makes It Special: {sv}")
            if inventory_row:
                iv = self.get_cell_value(sheet, inventory_row, col)
                if iv:
                    description_parts.append(f"Inventory In House: {iv}")
            if landing_row:
                lv = self.get_cell_value(sheet, landing_row, col)
                if lv:
                    description_parts.append(f"Landing Page: {lv}")

            # Featured Products (scan for "Featured Product N" labels in B column)
            featured_products = []
            for row in range(1, min(sheet.max_row + 1, 50)):
                lbl_b = self.get_cell_value(sheet, row, 2)  # Column B
                if lbl_b and 'featured product' in lbl_b.lower() and 'url' not in lbl_b.lower():
                    product_val = self.get_cell_value(sheet, row, col)
                    url_val = self.get_cell_value(sheet, row + 1, col)
                    if product_val and product_val.upper() != 'NO ADDITIONAL PRODUCTS':
                        entry = f"  - {product_val}"
                        if url_val:
                            entry += f"\n    URL: {url_val}"
                        featured_products.append(entry)

            if featured_products:
                description_parts.append("Featured Products:")
                description_parts.extend(featured_products)

            # Read SMS copy for optional [SMS] task (Email Check section not included in email description)
            sms_content_row = self._scan_for_label(sheet, 'SMS', 44, min(sheet.max_row, 50))
            sms_val = None
            if sms_content_row:
                sms_val = self.get_cell_value(sheet, sms_content_row, col)

            # Clean up description - remove empty sections
            final_parts = []
            for part in description_parts:
                if part.startswith('---') and final_parts and final_parts[-1].startswith('---'):
                    final_parts.pop()  # Remove consecutive section headers
                final_parts.append(part)

            task_description = '\n'.join(final_parts)

            # Build tags
            tags = [self.brand, week_num, 'Email Brief']
            if campaign_type:
                tags.append(campaign_type)

            task = {
                'Task Name': task_name,
                'Task Description': task_description,
                'Due Date': due_date or '',
                'Start Date': start_date or '',
                'Priority': priority,
                'Status': 'Open',
                'Tags': ', '.join(tags),
            }

            self.tasks.append(task)
            count += 1

            # Split SMS into separate task when campaign has SMS
            if 'SMS' in assets or sms_val:
                sms_task_name = f"[SMS] {task_name}"
                sms_desc_parts = [
                    "== SMS BRIEF ==",
                    f"Send Date: {start_date or 'TBD'} ({day_of_week or ''}) {time_of_send or ''}",
                    f"Campaign Type: {campaign_type or 'TBD'}",
                ]
                if row_overview:
                    overview = self.get_cell_value(sheet, row_overview, col)
                    if overview:
                        sms_desc_parts.append(f"Overview: {overview}")
                landing_row = find_row('Landing Page')
                if landing_row:
                    lv = self.get_cell_value(sheet, landing_row, col)
                    if lv:
                        sms_desc_parts.append(f"Landing Page: {lv}")
                sms_desc_parts.append(f"SMS Copy: {sms_val or '(to be added)'}")
                sms_tags = [t if t != 'Email Brief' else 'SMS Brief' for t in tags] + ['SMS']
                sms_task = {
                    'Task Name': sms_task_name,
                    'Task Description': '\n'.join(sms_desc_parts),
                    'Due Date': due_date or '',
                    'Start Date': start_date or '',
                    'Priority': priority,
                    'Status': 'Open',
                    'Tags': ', '.join(sms_tags),
                }
                self.tasks.append(sms_task)
                count += 1

        return count

    def get_available_weekly_sheets(self) -> List[str]:
        """
        Get list of available weekly sheet names in the workbook.

        Returns:
            List of sheet names that match weekly pattern
        """
        weekly_pattern = re.compile(r'(^Wk\d+$|_wk\d+)', re.IGNORECASE)
        skip_sheets = {'Product Launch Calendar', 'Content Calendar', 'Template',
                        'Sheet3', 'Marketing Pipeline'}
        
        available_sheets = []
        for sheet_name in self.workbook.sheetnames:
            if sheet_name in skip_sheets:
                continue
            if weekly_pattern.search(sheet_name):
                available_sheets.append(sheet_name)
        
        return sorted(available_sheets)

    def find_and_extract_all_weekly_sheets(self, selected_weeks: Optional[List[str]] = None) -> int:
        """
        Find all weekly sheets (Wk*, wk*, *_wk*) and extract campaign tasks.
        Prioritizes standard Wk# sheets (the email brief tabs).

        Args:
            selected_weeks: Optional list of specific week sheet names to process.
                          If None, processes all weekly sheets.

        Returns:
            Total number of campaign tasks extracted
        """
        total_count = 0
        # Match sheets like "Wk6", "Wk7", "Wk8", "PB_wk2_12", "TGW_wk2_12"
        weekly_pattern = re.compile(r'(^Wk\d+$|_wk\d+)', re.IGNORECASE)
        skip_sheets = {'Product Launch Calendar', 'Content Calendar', 'Template',
                        'Sheet3', 'Marketing Pipeline'}

        for sheet_name in self.workbook.sheetnames:
            if sheet_name in skip_sheets:
                continue
            if weekly_pattern.search(sheet_name):
                # If selected_weeks provided, only process those sheets
                if selected_weeks is not None and sheet_name not in selected_weeks:
                    continue
                    
                count = self.extract_campaign_tasks_from_sheet(sheet_name)
                if count > 0:
                    total_count += count
                    self.stats['sheets_processed'] += 1
                    print(f"Extracted {count} email brief tasks from '{sheet_name}'")

        self.stats['campaign_tasks'] = total_count
        return total_count

    def write_csv(self) -> bool:
        """
        Write tasks to ClickUp CSV format.

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            with open(self.output_file, 'w', newline='', encoding='utf-8-sig') as csvfile:
                fieldnames = [
                    'Task Name',
                    'Task Description',
                    'Due Date',
                    'Start Date',
                    'Priority',
                    'Status',
                    'Tags',
                ]

                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()

                for task in self.tasks:
                    writer.writerow(task)

            print(f"Wrote {len(self.tasks)} tasks to {self.output_file}")
            return True

        except Exception as e:
            print(f"Error writing CSV: {e}", file=sys.stderr)
            return False

    def print_summary(self):
        """Print conversion summary statistics."""
        print("\n" + "=" * 60)
        print("CONVERSION SUMMARY")
        print("=" * 60)
        print(f"Brand:                {self.brand}")
        print(f"Excel File:           {self.excel_file}")
        print(f"Output CSV:           {self.output_file}")
        print(f"Email Brief Tasks:    {self.stats['campaign_tasks']}")
        print(f"Weekly Sheets:        {self.stats['sheets_processed']}")
        print(f"Total Tasks:          {len(self.tasks)}")
        print("=" * 60 + "\n")

    def convert(self, selected_weeks: Optional[List[str]] = None) -> bool:
        """
        Run full conversion process.
        Focuses exclusively on Wk# tabs (email briefs) — the tasks Angie enters manually.
        Content Calendar, Product Launch Calendar, and Marketing Pipeline are skipped.

        Args:
            selected_weeks: Optional list of specific week sheet names to process.
                          If None, processes all weekly sheets.

        Returns:
            bool: True if successful, False otherwise
        """
        if not self.load_workbook_safe():
            return False

        # Extract email brief tasks from selected or all weekly sheets
        self.find_and_extract_all_weekly_sheets(selected_weeks)

        # Write to CSV
        if not self.write_csv():
            return False

        self.print_summary()
        return True


def main():
    """Main entry point."""
    if len(sys.argv) != 4:
        print("Usage: python dtc_to_clickup.py <excel_file> <brand> <output_csv>")
        print("  excel_file: Path to input DTC Calendar Excel file")
        print("  brand: Brand name (PB or TGW)")
        print("  output_csv: Path to output CSV file for ClickUp import")
        sys.exit(1)

    excel_file = sys.argv[1]
    brand = sys.argv[2]
    output_csv = sys.argv[3]

    converter = DTCtoClickUpConverter(excel_file, brand, output_csv)

    if not converter.convert():
        sys.exit(1)

    sys.exit(0)


if __name__ == '__main__':
    main()
