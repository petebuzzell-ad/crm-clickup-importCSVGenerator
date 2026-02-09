#!/usr/bin/env python3
"""
DTC Calendar to ClickUp CSV Converter
Converts DTC Calendar Excel files for Arcadia Digital brands (PB, TGW) into ClickUp-importable CSV format.
"""

import sys
import csv
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class DTCtoClickUpConverter:
    """Converts DTC Calendar Excel files to ClickUp CSV format."""

    # Priority mapping for campaign types
    CAMPAIGN_TYPE_PRIORITY = {
        'Product Launches': 'High',
        'Promotions': 'High',
        'Story Telling': 'Normal',
        'Brand Moments': 'Normal',
        'Problem Solving': 'Normal',
        'Promo': 'High',
    }

    # Priority mapping for product launch priorities
    PRODUCT_PRIORITY_MAP = {
        'A': 'Urgent',
        'B': 'High',
        'C': 'Normal',
        'D': 'Low',
    }

    def __init__(self, excel_file: str, brand: str, output_file: str):
        """
        Initialize converter.

        Args:
            excel_file: Path to input Excel file
            brand: Brand name (PB or TGW)
            output_file: Path to output CSV file
        """
        self.excel_file = excel_file
        self.brand = brand
        self.output_file = output_file
        self.workbook = None
        self.tasks: List[Dict] = []
        self.stats = {
            'product_launches': 0,
            'campaign_tasks': 0,
            'sheets_processed': 0,
        }

    def load_workbook_safe(self) -> bool:
        """
        Load Excel workbook with error handling.

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            self.workbook = load_workbook(self.excel_file, data_only=True)
            print(f"Loaded workbook: {self.excel_file}")
            return True
        except FileNotFoundError:
            print(f"Error: File not found - {self.excel_file}", file=sys.stderr)
            return False
        except Exception as e:
            print(f"Error loading workbook: {e}", file=sys.stderr)
            return False

    def get_cell_value(self, sheet, row: int, col: int) -> Optional[str]:
        """
        Get cell value safely, handling None and converting to string.

        Args:
            sheet: openpyxl worksheet
            row: Row number (1-indexed)
            col: Column number (1-indexed)

        Returns:
            Cleaned string value or None
        """
        cell = sheet.cell(row, col)
        value = cell.value
        if value is None:
            return None
        # Convert to string and strip whitespace
        value_str = str(value).strip()
        return value_str if value_str else None

    def parse_date(self, date_str: str) -> Optional[str]:
        """
        Parse various date formats and return MM/DD/YYYY format.

        Args:
            date_str: Date string in various formats

        Returns:
            Formatted date string MM/DD/YYYY or None
        """
        if not date_str:
            return None

        date_str = str(date_str).strip()

        # Try common formats
        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%m/%d/%Y',
            '%m/%d/%Y %H:%M',
            '%m/%d/%Y %I:%M %p',
            '%Y-%m-%d',
            '%m-%d-%Y',
        ]

        for fmt in formats:
            try:
                dt = datetime.strptime(date_str, fmt)
                return dt.strftime('%m/%d/%Y')
            except ValueError:
                continue

        # If datetime object, format it
        if isinstance(date_str, datetime):
            return date_str.strftime('%m/%d/%Y')

        return None

    def due_date_two_weeks_before(self, send_date_str: str) -> Optional[str]:
        """
        Given a date string (MM/DD/YYYY), return the date 14 days earlier in the same format.

        Args:
            send_date_str: Date string in MM/DD/YYYY format

        Returns:
            Formatted date string MM/DD/YYYY or None if parsing fails
        """
        if not send_date_str:
            return None
        parsed = self.parse_date(send_date_str)
        if not parsed:
            return None
        try:
            dt = datetime.strptime(parsed, '%m/%d/%Y')
            earlier = dt - timedelta(days=14)
            return earlier.strftime('%m/%d/%Y')
        except ValueError:
            return None

    def parse_due_date_from_header(self, date_str: str) -> Optional[str]:
        """
        Parse due date from header like 'DUE 2/3 10 AM CT'.

        Args:
            date_str: Header string containing due date

        Returns:
            Formatted date string MM/DD/YYYY or None
        """
        if not date_str:
            return None

        # Extract date pattern like "2/3" or "2/3/24"
        match = re.search(r'(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?', date_str)
        if match:
            month, day = match.group(1), match.group(2)
            year = match.group(3) if match.group(3) else str(datetime.now().year)

            # Handle 2-digit year
            if len(year) == 2:
                year = f"20{year}"

            try:
                dt = datetime(int(year), int(month), int(day))
                return dt.strftime('%m/%d/%Y')
            except ValueError:
                return None

        return None

    def clean_text(self, text: str) -> str:
        """
        Clean text: strip whitespace and normalize line breaks.

        Args:
            text: Text to clean

        Returns:
            Cleaned text
        """
        if not text:
            return ''
        return re.sub(r'\s+', ' ', str(text).strip())

    def normalize_multiline_text(self, text: str) -> str:
        """
        Normalize multiline text for CSV.

        Args:
            text: Text potentially containing line breaks

        Returns:
            Text with normalized line breaks
        """
        if not text:
            return ''
        # Replace multiple spaces with single space, preserve intentional breaks
        return re.sub(r'  +', ' ', str(text).strip())

    def extract_product_launch_tasks(self) -> int:
        """
        Extract product launch tasks from 'Product Launch Calendar' sheet.
        Handles both PB format (6 cols) and TGW format (8 cols with SKET Task, PAR Date, PO#).
        """
        sheet_name = 'Product Launch Calendar'

        if sheet_name not in self.workbook.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found")
            return 0

        sheet = self.workbook[sheet_name]
        count = 0

        # Auto-detect column layout from header row
        headers = {}
        for col in range(1, min(sheet.max_column + 1, 15)):
            val = self.get_cell_value(sheet, 1, col)
            if val:
                headers[val.lower().rstrip(':')] = col

        # Map to column indices (handle both PB and TGW formats)
        col_subcat = headers.get('subcat(s)', headers.get('subcat', 1))
        col_desc = headers.get('description', 2)
        col_launch = headers.get('launch date', 3)
        col_sport = headers.get('sport', None)
        col_priority = headers.get('priority (a-d)', headers.get('priority', 5))
        col_notes = headers.get('notes', 6)
        col_po = headers.get('po #', headers.get('po', None))
        col_sket = headers.get('sket task', headers.get('sket', None))

        for row in range(2, sheet.max_row + 1):
            description = self.get_cell_value(sheet, row, col_desc)
            launch_date_val = sheet.cell(row, col_launch).value

            if not description and not launch_date_val:
                continue

            launch_date = self.parse_date(str(launch_date_val)) if launch_date_val else None
            if not launch_date or not description:
                continue

            subcat = self.get_cell_value(sheet, row, col_subcat)
            priority = self.get_cell_value(sheet, row, col_priority)
            notes = self.get_cell_value(sheet, row, col_notes) if col_notes else None
            sport = self.get_cell_value(sheet, row, col_sport) if col_sport else None

            priority_mapped = self.PRODUCT_PRIORITY_MAP.get(priority, 'Normal') if priority else 'Normal'

            tags = [self.brand, 'Product Launch']
            if sport:
                tags.append(sport)

            desc_parts = []
            if subcat:
                desc_parts.append(f"SKU/Subcat: {subcat}")
            if col_sket:
                sket = self.get_cell_value(sheet, row, col_sket)
                if sket:
                    desc_parts.append(f"SKET Task: {sket}")
            if col_po:
                po = self.get_cell_value(sheet, row, col_po)
                if po:
                    desc_parts.append(f"PO #: {po}")
            if notes:
                desc_parts.append(f"Notes: {notes}")

            task = {
                'Task Name': description,
                'Task Description': '\n'.join(desc_parts),
                'Due Date': launch_date,
                'Start Date': '',
                'Priority': priority_mapped,
                'Status': 'Open',
                'Tags': ', '.join(tags),
            }

            self.tasks.append(task)
            count += 1

        self.stats['product_launches'] = count
        print(f"Extracted {count} product launch tasks from '{sheet_name}'")
        return count

    def _find_label_in_row(self, sheet, row: int) -> Optional[str]:
        """Check columns B and C for a label value (weekly sheets use both)."""
        for col in [2, 3]:  # B, C
            val = self.get_cell_value(sheet, row, col)
            if val:
                return val
        return None

    def _scan_for_label(self, sheet, keyword: str, start_row: int, end_row: int) -> Optional[int]:
        """Find the row containing a label keyword (case-insensitive) in columns B or C."""
        keyword_lower = keyword.lower()
        for row in range(start_row, end_row + 1):
            for col in [2, 3]:
                val = self.get_cell_value(sheet, row, col)
                if val and keyword_lower in val.lower():
                    return row
        return None

    def extract_campaign_tasks_from_sheet(self, sheet_name: str) -> int:
        """
        Extract campaign/email-brief tasks from a weekly sheet (Wk6, Wk7, etc.).
        Each column (D onward) with a send date represents one email brief = one ClickUp task.
        """
        if sheet_name not in self.workbook.sheetnames:
            return 0

        sheet = self.workbook[sheet_name]
        count = 0

        # Extract week number from sheet name
        match = re.search(r'[Ww]k(\d+)', sheet_name)
        week_num = f"Week {match.group(1)}" if match else sheet_name

        # Fallback due date from D1 (e.g., "DUE 2/3 10 AM CT") when send date is missing
        due_date_header = self.get_cell_value(sheet, 1, 4)  # D1
        due_date_fallback = self.parse_due_date_from_header(due_date_header) if due_date_header else None

        # Build a row-label index so we can handle varying row positions
        # across sheets. Scan rows 1-50 for known labels.
        label_rows = {}
        for row in range(1, min(sheet.max_row + 1, 55)):
            for col in [2, 3]:
                val = self.get_cell_value(sheet, row, col)
                if val:
                    label_rows[row] = val

        # Find key rows by label content
        def find_row(keyword, default=None):
            for r, lbl in label_rows.items():
                if keyword.lower() in lbl.lower():
                    return r
            return default

        row_date_of_send = find_row('Date of Send', 4)
        row_time_of_send = find_row('Time of Send', 6)
        row_campaign_type = find_row('Campaign Type', 7)
        row_campaign_name = find_row('Campaign Name', 8)
        row_overview = find_row('Overview', 9)

        # Find campaign columns: any column with a date value in the Date of Send row
        campaign_cols = []
        for col in range(4, min(sheet.max_column + 1, 20)):
            val = sheet.cell(row_date_of_send, col).value
            if val is not None:
                campaign_cols.append(col)

        if not campaign_cols:
            return 0

        # Process each campaign column = one email brief
        for col in campaign_cols:
            # Date of Send = start date for this email
            start_date_val = sheet.cell(row_date_of_send, col).value
            start_date = self.parse_date(str(start_date_val)) if start_date_val else None

            # Due date = 2 weeks before send date; fallback to header when send date missing
            due_date = self.due_date_two_weeks_before(start_date) if start_date else due_date_fallback

            # Day of week
            day_of_week_row = row_date_of_send + 1
            day_of_week = self.get_cell_value(sheet, day_of_week_row, col)

            # Time of Send
            time_of_send = self.get_cell_value(sheet, row_time_of_send, col)

            # Campaign Type
            campaign_type = self.get_cell_value(sheet, row_campaign_type, col)

            # Campaign Name
            campaign_name = self.get_cell_value(sheet, row_campaign_name, col)

            if not campaign_name:
                continue

            # Build task name: [Week] [Send Date Day] - [Campaign Type] - [Campaign Name]
            date_label = start_date or ''
            if day_of_week:
                date_label = f"{day_of_week} {start_date}" if start_date else day_of_week
            task_name = f"[{week_num}] "
            if campaign_type:
                task_name += f"{campaign_type}: "
            task_name += campaign_name

            # Map priority based on campaign type
            priority = self.CAMPAIGN_TYPE_PRIORITY.get(campaign_type, 'Normal') if campaign_type else 'Normal'

            # --- Build rich description with all email brief details ---
            description_parts = []

            # Header info
            description_parts.append(f"== EMAIL BRIEF ==")
            description_parts.append(f"Send Date: {start_date or 'TBD'} ({day_of_week or ''}) {time_of_send or ''}")
            description_parts.append(f"Campaign Type: {campaign_type or 'TBD'}")

            # Overview
            if row_overview:
                overview = self.get_cell_value(sheet, row_overview, col)
                if overview:
                    description_parts.append(f"Overview: {overview}")

            # Assets requested (Email, SMS, Site Banner) - scan for these labels
            assets = []
            for row in range(row_overview + 1 if row_overview else 10, row_overview + 6 if row_overview else 16):
                lbl = self._find_label_in_row(sheet, row)
                if lbl:
                    lbl_lower = lbl.lower().strip()
                    val = self.get_cell_value(sheet, row, col)
                    if lbl_lower == 'email' and val and val.lower() == 'yes':
                        assets.append('Email')
                    elif lbl_lower == 'sms' and val and val.lower() == 'yes':
                        assets.append('SMS')
                    elif 'site banner' in lbl_lower and val and val.lower() == 'yes':
                        assets.append('Site Banner')
            if assets:
                description_parts.append(f"Assets Needed: {', '.join(assets)}")

            # Promo details - find rows labeled Promo, Offer, Coupon Code
            promo_row = find_row('Promo')
            offer_row = find_row('Offer')
            coupon_row = find_row('Coupon Code') or find_row('Coupon')

            promo_parts = []
            if promo_row:
                pv = self.get_cell_value(sheet, promo_row, col)
                if pv:
                    promo_parts.append(f"Promo: {pv}")
            if offer_row:
                ov = self.get_cell_value(sheet, offer_row, col)
                if ov:
                    promo_parts.append(f"Offer: {ov}")
            if coupon_row:
                cv = self.get_cell_value(sheet, coupon_row, col)
                if cv:
                    promo_parts.append(f"Code: {cv}")
            if promo_parts:
                description_parts.append(f"--- Promotion ---")
                description_parts.extend(promo_parts)

            # Merchandising section - DAM, Hero, Landing Page, Featured Products
            description_parts.append(f"--- Merchandising ---")

            dam_row = find_row('DAM Assets')
            hero_row = find_row('Hero Product URL')
            special_row = find_row('makes this product special')
            inventory_row = find_row('Inventory In House')
            landing_row = find_row('Landing Page')

            if dam_row:
                dv = self.get_cell_value(sheet, dam_row, col)
                if dv:
                    description_parts.append(f"DAM Assets: {dv}")
            if hero_row:
                hv = self.get_cell_value(sheet, hero_row, col)
                if hv:
                    description_parts.append(f"Hero Product URL: {hv}")
            if special_row:
                sv = self.get_cell_value(sheet, special_row, col)
                if sv:
                    description_parts.append(f"What Makes It Special: {sv}")
            if inventory_row:
                iv = self.get_cell_value(sheet, inventory_row, col)
                if iv:
                    description_parts.append(f"Inventory In House: {iv}")
            if landing_row:
                lv = self.get_cell_value(sheet, landing_row, col)
                if lv:
                    description_parts.append(f"Landing Page: {lv}")

            # Featured Products (scan for "Featured Product N" labels in B column)
            featured_products = []
            for row in range(1, min(sheet.max_row + 1, 50)):
                lbl_b = self.get_cell_value(sheet, row, 2)  # Column B
                if lbl_b and 'featured product' in lbl_b.lower() and 'url' not in lbl_b.lower():
                    product_val = self.get_cell_value(sheet, row, col)
                    url_val = self.get_cell_value(sheet, row + 1, col)
                    if product_val and product_val.upper() != 'NO ADDITIONAL PRODUCTS':
                        entry = f"  - {product_val}"
                        if url_val:
                            entry += f"\n    URL: {url_val}"
                        featured_products.append(entry)

            if featured_products:
                description_parts.append("Featured Products:")
                description_parts.extend(featured_products)

            # Read SMS copy for optional [SMS] task (Email Check section not included in email description)
            sms_content_row = self._scan_for_label(sheet, 'SMS', 44, min(sheet.max_row, 50))
            sms_val = None
            if sms_content_row:
                sms_val = self.get_cell_value(sheet, sms_content_row, col)

            # Clean up description - remove empty sections
            final_parts = []
            for part in description_parts:
                if part.startswith('---') and final_parts and final_parts[-1].startswith('---'):
                    final_parts.pop()  # Remove consecutive section headers
                final_parts.append(part)

            task_description = '\n'.join(final_parts)

            # Build tags
            tags = [self.brand, week_num, 'Email Brief']
            if campaign_type:
                tags.append(campaign_type)

            task = {
                'Task Name': task_name,
                'Task Description': task_description,
                'Due Date': due_date or '',
                'Start Date': start_date or '',
                'Priority': priority,
                'Status': 'Open',
                'Tags': ', '.join(tags),
            }

            self.tasks.append(task)
            count += 1

            # Split SMS into separate task when campaign has SMS
            if 'SMS' in assets or sms_val:
                sms_task_name = f"[SMS] {task_name}"
                sms_desc_parts = [
                    "== SMS BRIEF ==",
                    f"Send Date: {start_date or 'TBD'} ({day_of_week or ''}) {time_of_send or ''}",
                    f"Campaign Type: {campaign_type or 'TBD'}",
                ]
                if row_overview:
                    overview = self.get_cell_value(sheet, row_overview, col)
                    if overview:
                        sms_desc_parts.append(f"Overview: {overview}")
                landing_row = find_row('Landing Page')
                if landing_row:
                    lv = self.get_cell_value(sheet, landing_row, col)
                    if lv:
                        sms_desc_parts.append(f"Landing Page: {lv}")
                sms_desc_parts.append(f"SMS Copy: {sms_val or '(to be added)'}")
                sms_tags = [t if t != 'Email Brief' else 'SMS Brief' for t in tags] + ['SMS']
                sms_task = {
                    'Task Name': sms_task_name,
                    'Task Description': '\n'.join(sms_desc_parts),
                    'Due Date': due_date or '',
                    'Start Date': start_date or '',
                    'Priority': priority,
                    'Status': 'Open',
                    'Tags': ', '.join(sms_tags),
                }
                self.tasks.append(sms_task)
                count += 1

        return count

    def find_and_extract_all_weekly_sheets(self) -> int:
        """
        Find all weekly sheets (Wk*, wk*, *_wk*) and extract campaign tasks.
        Prioritizes standard Wk# sheets (the email brief tabs).

        Returns:
            Total number of campaign tasks extracted
        """
        total_count = 0
        # Match sheets like "Wk6", "Wk7", "Wk8", "PB_wk2_12", "TGW_wk2_12"
        weekly_pattern = re.compile(r'(^Wk\d+$|_wk\d+)', re.IGNORECASE)
        skip_sheets = {'Product Launch Calendar', 'Content Calendar', 'Template',
                        'Sheet3', 'Marketing Pipeline'}

        for sheet_name in self.workbook.sheetnames:
            if sheet_name in skip_sheets:
                continue
            if weekly_pattern.search(sheet_name):
                count = self.extract_campaign_tasks_from_sheet(sheet_name)
                if count > 0:
                    total_count += count
                    self.stats['sheets_processed'] += 1
                    print(f"Extracted {count} email brief tasks from '{sheet_name}'")

        self.stats['campaign_tasks'] = total_count
        return total_count

    def write_csv(self) -> bool:
        """
        Write tasks to ClickUp CSV format.

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            with open(self.output_file, 'w', newline='', encoding='utf-8-sig') as csvfile:
                fieldnames = [
                    'Task Name',
                    'Task Description',
                    'Due Date',
                    'Start Date',
                    'Priority',
                    'Status',
                    'Tags',
                ]

                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()

                for task in self.tasks:
                    writer.writerow(task)

            print(f"Wrote {len(self.tasks)} tasks to {self.output_file}")
            return True

        except Exception as e:
            print(f"Error writing CSV: {e}", file=sys.stderr)
            return False

    def print_summary(self):
        """Print conversion summary statistics."""
        print("\n" + "=" * 60)
        print("CONVERSION SUMMARY")
        print("=" * 60)
        print(f"Brand:                {self.brand}")
        print(f"Excel File:           {self.excel_file}")
        print(f"Output CSV:           {self.output_file}")
        print(f"Email Brief Tasks:    {self.stats['campaign_tasks']}")
        print(f"Weekly Sheets:        {self.stats['sheets_processed']}")
        print(f"Total Tasks:          {len(self.tasks)}")
        print("=" * 60 + "\n")

    def convert(self) -> bool:
        """
        Run full conversion process.
        Focuses exclusively on Wk# tabs (email briefs) — the tasks Angie enters manually.
        Content Calendar, Product Launch Calendar, and Marketing Pipeline are skipped.
        """
        if not self.load_workbook_safe():
            return False

        # Extract email brief tasks from all weekly sheets only
        self.find_and_extract_all_weekly_sheets()

        # Write to CSV
        if not self.write_csv():
            return False

        self.print_summary()
        return True


def main():
    """Main entry point."""
    if len(sys.argv) != 4:
        print("Usage: python dtc_to_clickup.py <excel_file> <brand> <output_csv>")
        print("  excel_file: Path to input DTC Calendar Excel file")
        print("  brand: Brand name (PB or TGW)")
        print("  output_csv: Path to output CSV file for ClickUp import")
        sys.exit(1)

    excel_file = sys.argv[1]
    brand = sys.argv[2]
    output_csv = sys.argv[3]

    converter = DTCtoClickUpConverter(excel_file, brand, output_csv)

    if not converter.convert():
        sys.exit(1)

    sys.exit(0)


if __name__ == '__main__':
    main()
